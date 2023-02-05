import openpyxl
from pathlib import Path
import numpy as np 
import pandas as pd

class IDT_Gym:

    def __init__(self):
        """
        Initialize The IDT_Gyme environment
        """      
        self.load_IHM_Dataset()

    def load_IHM_Dataset(self):
        """
        Load The IHM dataset from excel into a pandas DataFrame
        """
        xlsx_file = Path('IHM', 'Events Matrix_V4.xlsx')
        wb_obj = openpyxl.load_workbook(xlsx_file)
        sheet = wb_obj["Flu_Records"]
        Pationet_Id = 'B'
        Cureent_s_col = 'C'
        Actions_col = 'D'
        Next_s_col = 'E'
        Cycle_col = 'F'
        Data_col_range = \
            ['G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V',
            'W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ',
            'AK','AL','AM','AN','AO','AP','AQ','AR']

        self.IHM_data = {}
        # Load IHM features
        for col in Data_col_range:
            Feature_Category = sheet[col][4].value
            Feature_name= sheet[col][6].value
            if not self.IHM_data.get(Feature_Category): 
                self.IHM_data[Feature_Category] = {}
            self.IHM_data[Feature_Category][Feature_name] = []
            for i in range(7,107):
                self.IHM_data[Feature_Category][Feature_name].append(sheet[col][i].value)
        # Load IHM Current state
        self.IHM_data['Current_s'] = []
        for i in range(7,107):
            self.IHM_data['Current_s'].append(sheet[Cureent_s_col][i].value)
        # Load IHM Action
        self.IHM_data['Actions'] = []
        for i in range(7,107):
            self.IHM_data['Actions'].append(sheet[Actions_col][i].value)
        # Load IHM Current state
        self.IHM_data['Next_s'] = []
        for i in range(7,107):
            self.IHM_data['Next_s'].append(sheet[Next_s_col][i].value)

            # 
            #     state = 

            #     self.inputs = {}

        # self.Number = np.asarray([])
        # self.Category = np.asarray([])
        # self.Group = np.asarray([])
        # self.UID = np.asarray([])
        # self.Name = np.asarray([])
        # self.ID = np.asarray([])
        # self.Cycle = np.asarray([])
        # self.Timee_stamp = np.asarray([])
        # self.Inputs = np.zeros((100,16))
        # self.Contexts = np.zeros((3,60))
        # self.Actions = np.zeros((7,60))
        # self.Outputs = np.zeros((9,60))
        # self.States = [Inputs,Context,Output]
        # for row in sheet.iter_rows(min_row=18, min_col=8, max_row=18, max_col=32):
        #     for cell in row:
        #         self.Number= np.append(self.Number,cell.value) 
        # for row in sheet.iter_rows(min_row=19, min_col=8, max_row=19, max_col=32):
        #     for cell in row:
        #         self.Category= np.append(self.Category,cell.value) 
        # for row in sheet.iter_rows(min_row=20, min_col=8, max_row=20, max_col=32):
        #     for cell in row:
        #         self.Group= np.append(self.Group,cell.value) 
        # for row in sheet.iter_rows(min_row=21, min_col=8, max_row=21, max_col=32):
        #     for cell in row:
        #         self.UID= np.append(self.UID,cell.value) 
        # for row in sheet.iter_rows(min_row=22, min_col=8, max_row=22, max_col=32):
        #     for cell in row:
        #         self.Name= np.append(self.Name,cell.value) 
        # for row in sheet.iter_rows(min_row=31, min_col=4, max_row=90, max_col=4):
        #     for cell in row:
        #         self.ID= np.append(self.ID,cell.value) 
        # for row in sheet.iter_rows(min_row=31, min_col=5, max_row=90, max_col=5):
        #     for cell in row:
        #         self.Cycle= np.append(self.Cycle,cell.value) 
        # for row in sheet.iter_rows(min_row=31, min_col=6, max_row=90, max_col=6):
        #     for cell in row:
        #         self.Timee_stamp= np.append(self.Timee_stamp,cell.value) 

        # i=0
        # for row in sheet.iter_rows(min_row=8, min_col=2, max_row=107, max_col=6):
        #     for j,cell in enumerate(row):
        #         if cell.value:
        #             try:
        #                 self.Inputs[i,j] = cell.value
        #                 print(self.Inputs[-1])
        #             except Exception as e:
        #                 print(cell.value)
        #     i+=1
        # df = pd.read_excel('IHM/Events Matrix_V4.xlsx', index_col=0,sheet_name="Flu_Records",skiprows=lambda x: x in range(6))  
        # print(df["Name"])
        # for i,row in enumerate(range(100)):
        #     for j,col in enumerate(range(15)):
        #         try:
        #             self.Inputs[i,j] = sheet[row][col].value
        #         except Exception as e:
        #             self.Inputs[i,j] = 0
        # # i = 0
        # # for row in sheet.iter_rows(min_row=31, min_col=8, max_row=90, max_col=13):
        # #     for j,cell in enumerate(row):
        # #         if cell.value:
        # #             self.Contexts[i,j] = cell.value
        # for row in sheet.iter_rows(min_row=8, min_col=23, max_row=107, max_col=38):
        #     for j,cell in enumerate(row):
        #         self.Outputs[i,j] = cell.value
        # for row in sheet.iter_rows(min_row=8, min_col=39, max_row=107, max_col=44):
        #     for j,cell in enumerate(row):
        #         self.Actions[i,j] = cell.value
        
    def Episodise(self):
        """
        Find identical states in a dataset, create a new dataset with all applicable self.Actionss/next states
        """
        Available_states = np.unique(self.IHM_data['Current_s'])
        current = np.asarray(self.IHM_data['Current_s'])
        self.episodes = {}
        for s in Available_states:
            idx = np.asarray(np.where(current== s)[0]).astype(np.int64)  
            self.episodes[s] = {'Action':[],'Next':[]}
            for id in idx:
                self.episodes[s]['Action'].append(self.IHM_data['Actions'][id])
                self.episodes[s]['Next'].append(self.IHM_data['Next_s'][id])
        

    def fetsh_record(self,current_s,action):
        """
        Giiven the passesd state-self.Actions pair, retrieve: next state, all applicable self.Actionss
        """
        self.unique_episodes = {}
        for state in self.episodes.keys():
            self.unique_episodes[state] = {'Action':[],'Next':[]}
            for i in range(len(self.episodes[state]["Action"])):
                if self.episodes[state]["Action"][i] not in self.unique_episodes[state]["Action"] and self.episodes[state]["Next"][i] not in self.unique_episodes[state]["Next"] :    
                    self.unique_episodes[state]["Action"].append(self.episodes[state]["Action"][i])
                    self.unique_episodes[state]["Next"].append(self.episodes[state]["Next"][i])

        next_s = self.unique_episodes[current_s]['Next']
        available_actions = self.unique_episodes[current_s]['Action']
        print("State,action pair taken: ",current_s,action)
        for i in range(len(next_s)):
            print("Available actions, next state pair: ",available_actions[i],next_s[i])
            

        return next_s,available_actions


    def execute_action(self):
        """
        Update the IHM with the executed self.Actions
        """
        pass

    def get_reward(self):
        """
        Given the passed state-self.Actions pair, retrieve: the reward from the IHM
        """
        pass

    def set_reward(self):
        """
        Given the passed state-self.Actions pair, retrieve: the reward from the IHM
        """
        pass
    def reset(self):
        '''
        Resets variables necessary to start next training episode.
        '''
if __name__ == '__main__':
    ENV = IDT_Gym()
    ENV.Episodise()
    next_s,available_actions = (ENV.fetsh_record("S3","A1"))
    # for i in range(len(ENV.Inputs)):
    #     print(ENV.Inputs[i])

